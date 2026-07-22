import base64
import json
import os
from datetime import datetime
from io import BytesIO

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image
from sqlalchemy import DateTime, Float, ForeignKey, LargeBinary, String, Text, create_engine, func, inspect, select, text
from sqlalchemy.orm import DeclarativeBase, Mapped, Session, joinedload, mapped_column, relationship

DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./data/voorraad.db")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql+psycopg://", 1)
elif DATABASE_URL.startswith("postgresql://") and "+psycopg" not in DATABASE_URL:
    DATABASE_URL = DATABASE_URL.replace("postgresql://", "postgresql+psycopg://", 1)

if DATABASE_URL.startswith("sqlite"):
    os.makedirs("data", exist_ok=True)
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
else:
    engine = create_engine(DATABASE_URL, pool_pre_ping=True)


class Base(DeclarativeBase):
    pass


class Product(Base):
    __tablename__ = "products"

    id: Mapped[int] = mapped_column(primary_key=True)
    name: Mapped[str] = mapped_column(String(200), index=True)
    article_number: Mapped[str] = mapped_column(String(100), default="", index=True)
    barcode: Mapped[str] = mapped_column(String(100), default="", index=True)
    category: Mapped[str] = mapped_column(String(100), default="")
    unit: Mapped[str] = mapped_column(String(40), default="stuks")
    location: Mapped[str] = mapped_column(String(100), default="")
    stock: Mapped[float] = mapped_column(Float, default=0)
    minimum_stock: Mapped[float] = mapped_column(Float, default=0)
    notes: Mapped[str] = mapped_column(Text, default="")
    photo_data: Mapped[str] = mapped_column(Text, default="")
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    mutations: Mapped[list["StockMutation"]] = relationship(
        back_populates="product", cascade="all, delete-orphan"
    )


class StockMutation(Base):
    __tablename__ = "stock_mutations"

    id: Mapped[int] = mapped_column(primary_key=True)
    product_id: Mapped[int] = mapped_column(ForeignKey("products.id"), index=True)
    change: Mapped[float] = mapped_column(Float)
    stock_after: Mapped[float] = mapped_column(Float)
    reason: Mapped[str] = mapped_column(String(200), default="")
    employee: Mapped[str] = mapped_column(String(100), default="")
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow, index=True)
    product: Mapped[Product] = relationship(back_populates="mutations")


class ExcelBackup(Base):
    __tablename__ = "excel_backups"

    id: Mapped[int] = mapped_column(primary_key=True)
    filename: Mapped[str] = mapped_column(String(200))
    file_data: Mapped[bytes] = mapped_column(LargeBinary)
    product_count: Mapped[int] = mapped_column(default=0)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow, index=True)


Base.metadata.create_all(engine)
with engine.begin() as connection:
    columns = {column["name"] for column in inspect(connection).get_columns("products")}
    if "photo_data" not in columns:
        connection.execute(text("ALTER TABLE products ADD COLUMN photo_data TEXT DEFAULT ''"))
    if "barcode" not in columns:
        connection.execute(text("ALTER TABLE products ADD COLUMN barcode VARCHAR(100) DEFAULT ''"))


PRODUCT_IMAGE_MAP = {}
try:
    with open("app/static/product_images.json", "r", encoding="utf-8") as image_file:
        PRODUCT_IMAGE_MAP = json.load(image_file)
except (OSError, ValueError):
    PRODUCT_IMAGE_MAP = {}

def product_image(product):
    if getattr(product, "photo_data", ""):
        return product.photo_data
    return PRODUCT_IMAGE_MAP.get(str(getattr(product, "barcode", "")), "/static/product-images/default.png")

app = FastAPI(title="Jan Bos Voorraad")
app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")
templates.env.globals["product_image"] = product_image


@app.get("/health")
def health():
    """Healthcheck that also verifies the configured database connection."""
    try:
        with engine.connect() as connection:
            connection.execute(text("SELECT 1"))
        return {
            "status": "ok",
            "database": engine.dialect.name,
            "persistent": engine.dialect.name == "postgresql",
        }
    except Exception as exc:
        return JSONResponse(
            status_code=503,
            content={"status": "error", "database": engine.dialect.name, "detail": str(exc)},
        )


def make_photo_data(upload: UploadFile | None) -> str:
    if not upload or not upload.filename:
        return ""
    raw = upload.file.read()
    if not raw:
        return ""
    if len(raw) > 8 * 1024 * 1024:
        raise HTTPException(400, "De foto is te groot. Gebruik maximaal 8 MB.")
    try:
        image = Image.open(BytesIO(raw)).convert("RGB")
        image.thumbnail((900, 900))
        output = BytesIO()
        image.save(output, format="JPEG", quality=78, optimize=True)
    except Exception as exc:
        raise HTTPException(400, "Dit bestand is geen geldige foto.") from exc
    return "data:image/jpeg;base64," + base64.b64encode(output.getvalue()).decode("ascii")


def clean_barcode(value: str) -> str:
    return "".join(value.strip().split())


def get_product_by_barcode(db: Session, barcode: str) -> Product | None:
    return db.scalar(select(Product).where(Product.barcode == clean_barcode(barcode)))


@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request, q: str = "", category: str = "", low: int = 0):
    with Session(engine) as db:
        stmt = select(Product)
        if q:
            like = f"%{q}%"
            stmt = stmt.where(
                (Product.name.ilike(like))
                | (Product.article_number.ilike(like))
                | (Product.barcode.ilike(like))
                | (Product.location.ilike(like))
            )
        if category:
            stmt = stmt.where(Product.category == category)
        if low:
            stmt = stmt.where(Product.stock <= Product.minimum_stock)
        products = list(db.scalars(stmt.order_by(Product.name)))
        categories = list(db.scalars(select(Product.category).where(Product.category != "").distinct().order_by(Product.category)))
        total = db.scalar(select(func.count(Product.id))) or 0
        low_count = db.scalar(select(func.count(Product.id)).where(Product.stock <= Product.minimum_stock)) or 0
        missing_photo_count = sum(1 for p in products if not p.photo_data and str(p.barcode) not in PRODUCT_IMAGE_MAP)
        popular_ids = [row[0] for row in db.execute(select(StockMutation.product_id, func.count(StockMutation.id)).group_by(StockMutation.product_id).order_by(func.count(StockMutation.id).desc()).limit(6))]
        popular = [db.get(Product, product_id) for product_id in popular_ids]
        mutations = list(db.scalars(select(StockMutation).options(joinedload(StockMutation.product)).order_by(StockMutation.created_at.desc()).limit(8)))
    return templates.TemplateResponse("index.html", {"request": request, "products": products, "popular": popular, "categories": categories, "q": q, "selected_category": category, "low": low, "total_products": total, "low_count": low_count, "missing_photo_count": missing_photo_count, "mutations": mutations})


@app.get("/inventarisatie", response_class=HTMLResponse)
def inventory_page(request: Request, barcode: str = ""):
    barcode = clean_barcode(barcode)
    product = None
    if barcode:
        with Session(engine) as db:
            product = get_product_by_barcode(db, barcode)
    return templates.TemplateResponse("inventory.html", {"request": request, "barcode": barcode, "product": product})


@app.get("/api/barcode/{barcode}")
def barcode_lookup(barcode: str):
    barcode = clean_barcode(barcode)
    with Session(engine) as db:
        product = get_product_by_barcode(db, barcode)
        if not product:
            return JSONResponse({"found": False, "barcode": barcode})
        return {"found": True, "barcode": barcode, "id": product.id, "name": product.name, "stock": product.stock, "unit": product.unit, "location": product.location}


@app.post("/inventory/create")
def inventory_create(
    barcode: str = Form(...), name: str = Form(...), article_number: str = Form(""),
    unit: str = Form("stuks"), location: str = Form(""), stock: float = Form(0),
    minimum_stock: float = Form(0), photo: UploadFile | None = File(None),
):
    barcode = clean_barcode(barcode)
    if not barcode or not name.strip():
        raise HTTPException(400, "Barcode en productnaam zijn verplicht")
    photo_data = make_photo_data(photo)
    with Session(engine) as db:
        if get_product_by_barcode(db, barcode):
            raise HTTPException(400, "Deze barcode is al gekoppeld")
        product = Product(name=name.strip(), article_number=article_number.strip(), barcode=barcode, unit=unit.strip() or "stuks", location=location.strip(), stock=stock, minimum_stock=minimum_stock, photo_data=photo_data)
        db.add(product)
        db.flush()
        db.add(StockMutation(product_id=product.id, change=stock, stock_after=stock, reason="Inventarisatie beginvoorraad", employee="Inventarisatie"))
        db.commit()
    return RedirectResponse("/inventarisatie?success=created", 303)


@app.post("/inventory/count/{product_id}")
def inventory_count(product_id: int, counted_stock: float = Form(...)):
    if counted_stock < 0:
        raise HTTPException(400, "Voorraad kan niet negatief zijn")
    with Session(engine) as db:
        product = db.get(Product, product_id)
        if not product:
            raise HTTPException(404, "Product niet gevonden")
        change = counted_stock - product.stock
        product.stock = counted_stock
        db.add(StockMutation(product_id=product.id, change=change, stock_after=counted_stock, reason="Voorraad geteld", employee="Inventarisatie"))
        db.commit()
    return RedirectResponse("/inventarisatie?success=counted", 303)


@app.get("/scan", response_class=HTMLResponse)
def scan_page(request: Request):
    return templates.TemplateResponse("scan.html", {"request": request})


@app.get("/product/{product_id}", response_class=HTMLResponse)
def product_page(request: Request, product_id: int, mode: str = ""):
    scan_mode = mode == "scan"
    with Session(engine) as db:
        product = db.get(Product, product_id)
        if not product:
            raise HTTPException(404, "Product niet gevonden")
        history = list(db.scalars(select(StockMutation).where(StockMutation.product_id == product_id).order_by(StockMutation.created_at.desc()).limit(10)))
    return templates.TemplateResponse("product.html", {"request": request, "p": product, "history": history, "scan_mode": scan_mode})


@app.post("/products")
def add_product(name: str = Form(...), article_number: str = Form(""), barcode: str = Form(""), category: str = Form(""), unit: str = Form("stuks"), location: str = Form(""), stock: float = Form(0), minimum_stock: float = Form(0), photo: UploadFile | None = File(None)):
    if not name.strip():
        raise HTTPException(400, "Productnaam ontbreekt")
    barcode = clean_barcode(barcode)
    photo_data = make_photo_data(photo)
    with Session(engine) as db:
        if barcode and get_product_by_barcode(db, barcode):
            raise HTTPException(400, "Deze barcode is al gekoppeld")
        product = Product(name=name.strip(), article_number=article_number.strip(), barcode=barcode, category=category.strip(), unit=unit.strip() or "stuks", location=location.strip(), stock=stock, minimum_stock=minimum_stock, photo_data=photo_data)
        db.add(product); db.flush()
        if stock:
            db.add(StockMutation(product_id=product.id, change=stock, stock_after=stock, reason="Beginvoorraad", employee="Systeem"))
        db.commit()
    return RedirectResponse("/", 303)


def change_stock(product_id: int, change: float, reason: str = "Snel geboekt", employee: str = ""):
    with Session(engine) as db:
        product = db.get(Product, product_id)
        if not product:
            raise HTTPException(404, "Product niet gevonden")
        if product.stock + change < 0:
            raise HTTPException(400, "Onvoldoende voorraad")
        product.stock += change
        db.add(StockMutation(product_id=product.id, change=change, stock_after=product.stock, reason=reason.strip(), employee=employee.strip()))
        db.commit()


@app.post("/products/{product_id}/quick")
def quick(product_id: int, change: float = Form(...), return_to: str = Form("/")):
    change_stock(product_id, change)
    return RedirectResponse(return_to if return_to.startswith("/") else "/", 303)


@app.post("/products/{product_id}/mutate")
def mutate(product_id: int, amount: float = Form(...), direction: str = Form(...), reason: str = Form(""), employee: str = Form(""), return_to: str = Form("/")):
    if amount <= 0:
        raise HTTPException(400, "Aantal moet groter zijn dan nul")
    change_stock(product_id, amount if direction == "in" else -amount, reason, employee)
    return RedirectResponse(return_to if return_to.startswith("/") else "/", 303)


@app.post("/products/{product_id}/edit")
def edit(product_id: int, name: str = Form(...), article_number: str = Form(""), barcode: str = Form(""), category: str = Form(""), unit: str = Form("stuks"), location: str = Form(""), minimum_stock: float = Form(0), photo: UploadFile | None = File(None), remove_photo: int = Form(0)):
    barcode = clean_barcode(barcode)
    with Session(engine) as db:
        product = db.get(Product, product_id)
        if not product:
            raise HTTPException(404, "Product niet gevonden")
        existing = get_product_by_barcode(db, barcode) if barcode else None
        if existing and existing.id != product_id:
            raise HTTPException(400, "Deze barcode is al gekoppeld")
        product.name = name.strip(); product.article_number = article_number.strip(); product.barcode = barcode
        product.category = category.strip(); product.unit = unit.strip() or "stuks"; product.location = location.strip(); product.minimum_stock = minimum_stock
        if remove_photo: product.photo_data = ""
        elif photo and photo.filename: product.photo_data = make_photo_data(photo)
        db.commit()
    return RedirectResponse(f"/product/{product_id}", 303)


@app.get("/fotos", response_class=HTMLResponse)
def photo_page(request: Request):
    with Session(engine) as db:
        missing = list(db.scalars(select(Product).where((Product.photo_data == "") | (Product.photo_data.is_(None))).order_by(Product.name)))
        product = missing[0] if missing else None
        total_missing = len(missing)
    return templates.TemplateResponse("photos.html", {"request": request, "p": product, "total_missing": total_missing})


@app.post("/products/{product_id}/photo")
def save_product_photo(product_id: int, photo: UploadFile = File(...), return_to: str = Form("/")):
    photo_data = make_photo_data(photo)
    with Session(engine) as db:
        product = db.get(Product, product_id)
        if not product:
            raise HTTPException(404, "Product niet gevonden")
        product.photo_data = photo_data
        db.commit()
    return RedirectResponse(return_to if return_to.startswith("/") else "/", 303)


@app.post("/products/{product_id}/delete")
def delete(product_id: int):
    with Session(engine) as db:
        product = db.get(Product, product_id)
        if product: db.delete(product); db.commit()
    return RedirectResponse("/", 303)


@app.get("/backups", response_class=HTMLResponse)
def backups_page(request: Request):
    with Session(engine) as db:
        backups = list(db.scalars(select(ExcelBackup).order_by(ExcelBackup.created_at.desc()).limit(50)))
    return templates.TemplateResponse("backups.html", {"request": request, "backups": backups})


def build_inventory_workbook(products, mutations):
    wb = Workbook()
    ws = wb.active
    ws.title = "Voorraad"
    ws.sheet_view.showGridLines = False
    navy = "173F5F"; blue = "24699F"; green = "C6EFCE"; green_text = "006100"
    red = "FFC7CE"; red_text = "9C0006"; orange = "FCE4D6"; orange_text = "9C5700"; light = "EAF1F5"
    thin = Side(style="thin", color="D9E2E8")

    ws.merge_cells("A1:I1")
    ws["A1"] = "JAN BOS VOORRAADOVERZICHT"
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32
    ws["A2"] = "Laatste update"; ws["B2"] = datetime.now().strftime("%d-%m-%Y %H:%M")
    ws["D2"] = "Aantal producten"; ws["E2"] = len(products)
    ws["G2"] = "Onder minimum"; ws["H2"] = sum(1 for p in products if p.stock < p.minimum_stock)
    for c in ("A2","D2","G2"):
        ws[c].font = Font(bold=True, color=navy)
    headers = ["Product", "Artikelcode", "Barcode", "Categorie", "Voorraad", "Eenheid", "Minimum", "Status", "Bijbestellen"]
    ws.append([]); ws.append(headers)
    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.alignment = Alignment(horizontal="center")
    for idx, product in enumerate(products, start=5):
        ws.append([product.name, product.article_number, product.barcode, product.category, product.stock, product.unit, product.minimum_stock, f'=IF(E{idx}<G{idx},"ROOD",IF(E{idx}=G{idx},"ORANJE","GROEN"))', f'=MAX(G{idx}-E{idx},0)'])
        for cell in ws[idx]:
            cell.border = Border(bottom=thin)
        ws[f"E{idx}"].number_format = '0.##'
        ws[f"G{idx}"].number_format = '0.##'
        ws[f"I{idx}"].number_format = '0.##'
    last=max(5, 4+len(products))
    ws.conditional_formatting.add(f"A5:I{last}", FormulaRule(formula=["$E5<$G5"], fill=PatternFill("solid", fgColor=red), font=Font(color=red_text)))
    ws.conditional_formatting.add(f"A5:I{last}", FormulaRule(formula=["$E5=$G5"], fill=PatternFill("solid", fgColor=orange), font=Font(color=orange_text)))
    ws.conditional_formatting.add(f"A5:I{last}", FormulaRule(formula=["$E5>$G5"], fill=PatternFill("solid", fgColor=green), font=Font(color=green_text)))
    widths=[34,16,20,18,12,14,12,14,14]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A5"; ws.auto_filter.ref=f"A4:I{last}"

    order = wb.create_sheet("Bijbestellen")
    order.sheet_view.showGridLines=False
    order.append(["Product", "Artikelcode", "Barcode", "Voorraad", "Minimum", "Te bestellen", "Eenheid"])
    for c in order[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor=navy)
    for product in products:
        if product.stock < product.minimum_stock:
            order.append([product.name, product.article_number, product.barcode, product.stock, product.minimum_stock, max(product.minimum_stock-product.stock,0), product.unit])
    for i,w in enumerate([34,16,20,12,12,14,14],1): order.column_dimensions[get_column_letter(i)].width=w
    order.freeze_panes="A2"; order.auto_filter.ref=f"A1:G{max(1,order.max_row)}"

    hist = wb.create_sheet("Mutaties")
    hist.sheet_view.showGridLines=False
    hist.append(["Datum", "Tijd", "Product", "Barcode", "Wijziging", "Voorraad na", "Reden", "Medewerker"])
    for c in hist[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor=navy)
    for m in mutations:
        hist.append([m.created_at.strftime("%d-%m-%Y"), m.created_at.strftime("%H:%M"), m.product.name if m.product else "", m.product.barcode if m.product else "", m.change, m.stock_after, m.reason, m.employee])
    for i,w in enumerate([13,10,34,20,12,14,28,18],1): hist.column_dimensions[get_column_letter(i)].width=w
    hist.freeze_panes="A2"; hist.auto_filter.ref=f"A1:H{max(1,hist.max_row)}"

    out=BytesIO(); wb.save(out); return out.getvalue()


@app.get("/export.xlsx")
def export_excel():
    with Session(engine) as db:
        products = list(db.scalars(select(Product).order_by(Product.name)))
        mutations = list(db.scalars(select(StockMutation).options(joinedload(StockMutation.product)).order_by(StockMutation.created_at.desc())))
        data = build_inventory_workbook(products, mutations)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = ExcelBackup(filename=f"Jan_Bos_Voorraad_backup_{stamp}.xlsx", file_data=data, product_count=len(products))
        db.add(backup); db.commit()
    return StreamingResponse(BytesIO(data), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="Jan_Bos_Voorraad.xlsx"'})


@app.get("/backups/{backup_id}/download")
def download_backup(backup_id: int):
    with Session(engine) as db:
        backup = db.get(ExcelBackup, backup_id)
        if not backup: raise HTTPException(404, "Back-up niet gevonden")
        data = backup.file_data; filename = backup.filename
    return StreamingResponse(BytesIO(data), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

@app.post("/api/products/{product_id}/quick")
def api_quick(product_id: int, change: float = Form(...)):
    with Session(engine) as db:
        product = db.get(Product, product_id)
        if not product:
            raise HTTPException(404, "Product niet gevonden")
        if product.stock + change < 0:
            raise HTTPException(400, "Onvoldoende voorraad")
        product.stock += change
        db.add(StockMutation(
            product_id=product.id,
            change=change,
            stock_after=product.stock,
            reason="Snel geboekt",
            employee="",
        ))
        db.commit()
        return {
            "ok": True,
            "stock": product.stock,
            "low": product.stock <= product.minimum_stock,
        }
